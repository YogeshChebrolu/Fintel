"""
groww_adapter.py — the ONLY code that understands Groww's P&L layout.

Groww's 'Trade Level' sheet is a sectioned report, not a flat table, so the parser
is a section state machine. Output is canonical objects only; nothing Groww-shaped
leaves this module.

Layout (observed):
  Title row        : 'P&L Statement for stocks from <start> TO <end>'   -> period
  'Summary' / 'P&L': Realised P&L, Unrealised P&L                        -> checksums
  'Charges'        : label -> value rows                                 -> Charges
  'Realised trades': header [Stock name, ISIN, Qty, Buy date, Buy price,
                     Buy value, Sell date, Sell price, Sell value,
                     Realised P&L, Remark]                               -> MatchedLot[]
  'Unrealised trades': header [..., Closing date, Closing price,
                     Closing value, Unrealised P&L, Remark]              -> OpenLot[]
  'Disclaimer'     : end
'Scrip Level' sheet is fully derivable from 'Trade Level' -> cross-check only.
"""
from __future__ import annotations

import hashlib
import re
from datetime import date, datetime
from decimal import Decimal

from openpyxl import load_workbook

from lib.types import (
    AssetClass, Charges, D, Execution, Instrument, MatchedLot, OpenLot,
    PortfolioState, Reconciliation, Segment, Side, SourceRef, money,
)

BROKER = "GROWW"
RECON_TOLERANCE = Decimal("0.10")   # absorb broker rounding noise only

_DATE_RE = re.compile(r"(\d{2}-\d{2}-\d{4})\s+TO\s+(\d{2}-\d{2}-\d{4})", re.I)

_CHARGE_MAP = {
    "Exchange Transaction Charges": "exchange_txn",
    "SEBI Charges": "sebi",
    "STT": "stt",
    "Stamp Duty": "stamp_duty",
    "IPFT Charges": "ipft",
    "Brokerage": "brokerage",
    "DP Charges": "dp",
    "Total GST": "gst",
    "Total": "total",
}


def _pdate(s) -> date:
    return datetime.strptime(str(s).strip(), "%d-%m-%Y").date()


def _hash(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _instrument(name: str, isin: str) -> Instrument:
    isin = (isin or "").strip()
    return Instrument(
        isin=isin,
        broker_symbol=str(name).strip(),
        asset_class=Instrument.asset_class_from_isin(isin),
        entity_key=isin or f"NAME::{str(name).strip()}",  # ISIN-keyed; fallback to name
    )


def _segment(remark, buy_d: date, sell_d: date) -> Segment:
    if remark and "intraday" in str(remark).lower():
        return Segment.INTRADAY
    if buy_d == sell_d:                       # same-day round trip => intraday
        return Segment.INTRADAY
    return Segment.DELIVERY


def parse(path: str) -> PortfolioState:
    file_hash = _hash(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Trade Level"]
    rows = [r for r in ws.iter_rows(values_only=True)]

    period_start = period_end = None
    realized_stated = unrealized_stated = Decimal(0)
    charges = Charges()
    matched: list[MatchedLot] = []
    opens: list[OpenLot] = []

    section = None  # None | 'charges' | 'realised' | 'unrealised'
    for idx, raw in enumerate(rows):
        if not raw:
            continue
        cells = list(raw)
        c0 = (cells[0] if cells else None)
        c0s = str(c0).strip() if c0 is not None else ""

        # ---- section markers -------------------------------------------
        if c0s.startswith("P&L Statement for stocks"):
            m = _DATE_RE.search(c0s)
            if m:
                period_start, period_end = _pdate(m.group(1)), _pdate(m.group(2))
            continue
        if c0s == "Realised P&L":
            realized_stated = D(cells[1]); continue
        if c0s == "Unrealised P&L":
            unrealized_stated = D(cells[1]); continue
        if c0s == "Charges":
            section = "charges"; continue
        if c0s in _CHARGE_MAP and section == "charges":
            setattr(charges, _CHARGE_MAP[c0s], money(cells[1])); continue
        if c0s == "Realised trades":
            section = "realised"; continue
        if c0s == "Unrealised trades":
            section = "unrealised"; continue
        if c0s.startswith("Disclaimer"):
            break
        if c0s == "Stock name":     # table header row; skip
            continue

        # ---- data rows -------------------------------------------------
        as_of = period_end or date.today()
        if section == "realised" and c0s and c0s != "Total" and len(cells) >= 10:
            name, isin, qty, bd, bp, bv, sd, sp, sv, pnl = cells[:10]
            remark = cells[10] if len(cells) > 10 else None
            if isin is None:        # 'Total' / blank guard
                continue
            inst = _instrument(name, isin)
            buy_d, sell_d = _pdate(bd), _pdate(sd)
            seg = _segment(remark, buy_d, sell_d)
            src = SourceRef(BROKER, file_hash, "Trade Level", idx + 1, as_of)
            buy = Execution(inst, Side.BUY, D(qty), money(bp), buy_d, seg, money(bv), src)
            sell = Execution(inst, Side.SELL, D(qty), money(sp), sell_d, seg, money(sv), src)
            matched.append(MatchedLot(
                instrument=inst, quantity=D(qty), buy=buy, sell=sell,
                realized_pnl=money(pnl), holding_days=(sell_d - buy_d).days,
                segment=seg, source=src,
            ))

        elif section == "unrealised" and c0s and c0s != "Total" and len(cells) >= 10:
            name, isin, qty, bd, bp, bv, cd, cp, cv, pnl = cells[:10]
            if isin is None:
                continue
            inst = _instrument(name, isin)
            buy_d = _pdate(bd)
            close_d = _pdate(cd)
            src = SourceRef(BROKER, file_hash, "Trade Level", idx + 1, close_d)
            opens.append(OpenLot(
                instrument=inst, quantity=D(qty), buy_date=buy_d,
                buy_price=money(bp), buy_value=money(bv),
                closing_price=money(cp), closing_value=money(cv),
                unrealized_pnl=money(pnl), holding_days=(close_d - buy_d).days,
                source=src,
            ))

    realized_computed = money(sum((l.realized_pnl for l in matched), Decimal(0)))
    unrealized_computed = money(sum((l.unrealized_pnl for l in opens), Decimal(0)))
    recon = Reconciliation(
        realized_stated=money(realized_stated), realized_computed=realized_computed,
        unrealized_stated=money(unrealized_stated), unrealized_computed=unrealized_computed,
        tolerance=RECON_TOLERANCE,
    )
    if not recon.realized_ok:
        recon.notes.append("Realised P&L sum does not match Summary; rows may be missing.")
    if not recon.unrealized_ok:
        recon.notes.append("Unrealised P&L sum does not match Summary.")

    return PortfolioState(
        account_ref=hashlib.sha256(b"client-code-placeholder").hexdigest()[:16],
        broker=BROKER,
        period_start=period_start, period_end=period_end,
        matched_lots=matched, open_lots=opens, charges=charges,
        realized_pnl=realized_computed, unrealized_pnl=unrealized_computed,
        reconciliation=recon, source_file_hash=file_hash,
    )


if __name__ == "__main__":
    import sys
    ps = parse(sys.argv[1])
    print(f"broker={ps.broker}  period={ps.period_start}..{ps.period_end}")
    print(f"matched lots = {len(ps.matched_lots)}   open lots = {len(ps.open_lots)}")
    print(f"distinct instruments = {len(ps.instruments())}")
    print(f"realised  stated={ps.reconciliation.realized_stated}  "
          f"computed={ps.reconciliation.realized_computed}  ok={ps.reconciliation.realized_ok}")
    print(f"unrealised stated={ps.reconciliation.unrealized_stated}  "
          f"computed={ps.reconciliation.unrealized_computed}  ok={ps.reconciliation.unrealized_ok}")
    print(f"charges total = {ps.charges.total}   reconciliation passed = {ps.reconciliation.passed}")
    intraday = sum(1 for l in ps.matched_lots if l.segment == Segment.INTRADAY)
    print(f"intraday matched lots = {intraday}  delivery = {len(ps.matched_lots) - intraday}")
    funds = sum(1 for i in ps.instruments().values() if i.asset_class == AssetClass.FUND)
    print(f"instruments by class: FUND={funds}  EQUITY={len(ps.instruments())-funds}")
